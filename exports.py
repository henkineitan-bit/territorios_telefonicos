"""
exports.py
----------
Funciones auxiliares para:
- Importar territorios/registros desde un archivo Excel (upsert)
- Exportar registros a Excel
- Exportar un territorio a PDF
- Exportar un territorio a PNG

Cada función devuelve un objeto BytesIO listo para mandar con send_file,
excepto importar_excel que devuelve un resumen (dict) de lo que hizo.
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from PIL import Image, ImageDraw, ImageFont


# ---------------------------------------------------------------------------
# Helpers para interpretar celdas de Excel de forma flexible
# ---------------------------------------------------------------------------

def _celda_a_texto(valor):
    """Convierte una celda de Excel a texto limpio, o None si está vacía."""
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _celda_a_bool01(valor):
    """
    Convierte una celda a 0 o 1, aceptando varias formas de escribir "sí/no":
    1, 0, "si", "sí", "no", "true", "false", vacío (-> 0).
    """
    if valor is None:
        return 0
    texto = str(valor).strip().lower()
    if texto in ("1", "si", "sí", "true", "x"):
        return 1
    return 0


def _celda_a_funcionan(valor):
    """
    Convierte una celda al valor de 'funcionan': None (sin verificar), 1 o 0.
    Deja en None cualquier celda vacía o que no se entienda con claridad.
    """
    if valor is None:
        return None
    texto = str(valor).strip().lower()
    if texto in ("1", "si", "sí", "true"):
        return 1
    if texto in ("0", "no", "false"):
        return 0
    return None


# ---------------------------------------------------------------------------
# Importar Excel (upsert de territorios + registros)
# ---------------------------------------------------------------------------

# Columnas esperadas en la primera fila (encabezado) del Excel.
# El orden no importa, se buscan por nombre.
COLUMNAS_ESPERADAS = [
    "numero", "direccion", "telefono", "observaciones",
    "no_llamar", "funcionan", "notas_internas",
]


def importar_excel(archivo, conn):
    """
    Lee un archivo Excel (file-like) y hace upsert en 'territorios' y 'registros'.

    - 'numero' y 'telefono' son obligatorios en cada fila; si faltan, la fila
      se cuenta como error y se saltea (no frena la importación completa).
    - El territorio se crea si no existe (por 'numero', que es UNIQUE).
    - El registro se inserta o actualiza según la clave única
      (territorio_id, direccion, telefono) que ya define el esquema.

    Devuelve un diccionario con contadores para mostrarle un resumen al usuario.
    """
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active

    filas = list(hoja.iter_rows(values_only=False))
    if not filas:
        return {"territorios_creados": 0, "registros_insertados": 0,
                "registros_actualizados": 0, "errores": 0, "detalle_errores": []}

    # Mapeamos nombre de columna -> índice, leyendo la fila de encabezado
    encabezado = [str(c.value).strip().lower() if c.value else "" for c in filas[0]]
    indice = {}
    for nombre in COLUMNAS_ESPERADAS:
        if nombre in encabezado:
            indice[nombre] = encabezado.index(nombre)

    if "numero" not in indice or "telefono" not in indice:
        raise ValueError(
            "El Excel debe tener al menos las columnas 'numero' y 'telefono' "
            "en la primera fila."
        )

    resumen = {
        "territorios_creados": 0,
        "registros_insertados": 0,
        "registros_actualizados": 0,
        "errores": 0,
        "detalle_errores": [],
    }

    def valor(fila, nombre):
        i = indice.get(nombre)
        return fila[i].value if i is not None and i < len(fila) else None

    for num_fila, fila in enumerate(filas[1:], start=2):  # fila 1 es encabezado
        numero = _celda_a_texto(valor(fila, "numero"))
        telefono = _celda_a_texto(valor(fila, "telefono"))

        if not numero or not telefono:
            resumen["errores"] += 1
            resumen["detalle_errores"].append(
                f"Fila {num_fila}: falta 'numero' o 'telefono', se saltea."
            )
            continue

        direccion = _celda_a_texto(valor(fila, "direccion"))
        observaciones = _celda_a_texto(valor(fila, "observaciones"))
        notas_internas = _celda_a_texto(valor(fila, "notas_internas"))
        no_llamar = _celda_a_bool01(valor(fila, "no_llamar"))
        funcionan = _celda_a_funcionan(valor(fila, "funcionan"))

        # --- Upsert de territorio por 'numero' ---
        territorio = conn.execute(
            "SELECT id FROM territorios WHERE numero = ?", (numero,)
        ).fetchone()
        if territorio is None:
            cur = conn.execute(
                "INSERT INTO territorios (numero, estado) VALUES (?, 'Disponible')",
                (numero,),
            )
            territorio_id = cur.lastrowid
            resumen["territorios_creados"] += 1
        else:
            territorio_id = territorio["id"]

        # --- Upsert de registro por (territorio_id, direccion, telefono) ---
        existente = conn.execute(
            """
            SELECT id FROM registros
            WHERE territorio_id = ?
              AND direccion IS ?
              AND telefono = ?
            """,
            (territorio_id, direccion, telefono),
        ).fetchone()

        if existente is None:
            conn.execute(
                """
                INSERT INTO registros
                    (territorio_id, direccion, telefono, observaciones,
                     no_llamar, funcionan, notas_internas)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (territorio_id, direccion, telefono, observaciones,
                 no_llamar, funcionan, notas_internas),
            )
            resumen["registros_insertados"] += 1
        else:
            conn.execute(
                """
                UPDATE registros
                SET observaciones = ?, no_llamar = ?, funcionan = ?, notas_internas = ?
                WHERE id = ?
                """,
                (observaciones, no_llamar, funcionan, notas_internas, existente["id"]),
            )
            resumen["registros_actualizados"] += 1

    conn.commit()
    return resumen


# ---------------------------------------------------------------------------
# Exportar a Excel
# ---------------------------------------------------------------------------

def generar_excel(registros):
    """
    Recibe una lista de filas (sqlite3.Row) con columnas de 'registros' +
    'numero_territorio', y devuelve un BytesIO con un .xlsx listo para descargar.
    """
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Registros"

    encabezados = [
        "numero_territorio", "direccion", "telefono", "observaciones",
        "no_llamar", "funcionan", "notas_internas",
    ]
    hoja.append(encabezados)

    for reg in registros:
        funcionan = reg["funcionan"]
        funcionan_texto = "" if funcionan is None else ("Si" if funcionan else "No")
        hoja.append([
            reg["numero_territorio"],
            reg["direccion"] or "",
            reg["telefono"],
            reg["observaciones"] or "",
            "Si" if reg["no_llamar"] else "No",
            funcionan_texto,
            reg["notas_internas"] or "",
        ])

    # Ancho de columnas prolijo, para que no quede todo apretado
    anchos = [16, 30, 16, 30, 10, 12, 30]
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[hoja.cell(row=1, column=i).column_letter].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Exportar territorio a PDF
# ---------------------------------------------------------------------------

def generar_pdf(territorio, registros):
    """
    Genera un PDF simple e imprimible con los datos del territorio y su
    tabla de registros telefónicos.
    """
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    ancho, alto = A4

    margen = 2 * cm
    y = alto - margen

    c.setFont("Helvetica-Bold", 16)
    c.drawString(margen, y, f"Territorio {territorio['numero']}")
    y -= 0.8 * cm

    c.setFont("Helvetica", 11)
    c.drawString(margen, y, f"Estado: {territorio['estado']}")
    y -= 1 * cm

    # Encabezados de la tabla
    columnas = ["Dirección", "Teléfono", "Observaciones", "No llamar", "Funciona"]
    anchos_col = [5 * cm, 3 * cm, 5 * cm, 2.2 * cm, 2.2 * cm]

    def dibujar_encabezado(y):
        c.setFont("Helvetica-Bold", 9)
        x = margen
        for texto, ancho_col in zip(columnas, anchos_col):
            c.drawString(x, y, texto)
            x += ancho_col
        return y - 0.5 * cm

    y = dibujar_encabezado(y)
    c.setFont("Helvetica", 9)

    for reg in registros:
        if y < margen:  # se acabó la hoja, arrancamos una nueva página
            c.showPage()
            y = alto - margen
            y = dibujar_encabezado(y)
            c.setFont("Helvetica", 9)

        if reg["funcionan"] is None:
            funciona_txt = "Sin verificar"
        elif reg["funcionan"]:
            funciona_txt = "Si"
        else:
            funciona_txt = "No"

        valores = [
            (reg["direccion"] or "—")[:28],
            reg["telefono"],
            (reg["observaciones"] or "—")[:28],
            "Si" if reg["no_llamar"] else "No",
            funciona_txt,
        ]

        x = margen
        for texto, ancho_col in zip(valores, anchos_col):
            c.drawString(x, y, str(texto))
            x += ancho_col

        y -= 0.5 * cm

    c.save()
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Exportar territorio a PNG
# ---------------------------------------------------------------------------

def generar_png(territorio, registros):
    """
    Genera una imagen PNG con una tabla simple del territorio, pensada para
    compartir rápido (ej. por WhatsApp) sin necesidad de abrir un PDF.
    """
    ancho_img = 900
    fila_alto = 30
    alto_img = 100 + fila_alto * (len(registros) + 1)

    img = Image.new("RGB", (ancho_img, max(alto_img, 150)), color="white")
    draw = ImageDraw.Draw(img)

    try:
        fuente_titulo = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20
        )
        fuente_normal = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14
        )
        fuente_negrita = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14
        )
    except OSError:
        # Intentar con fuentes de Windows
        try:
            fuente_titulo = ImageFont.truetype("arial.ttf", 20)
            fuente_normal = ImageFont.truetype("arial.ttf", 14)
            fuente_negrita = ImageFont.truetype("arialbd.ttf", 14)
        except OSError:
            # Último recurso: fuente por defecto de Pillow
            fuente_titulo = ImageFont.load_default()
            fuente_normal = ImageFont.load_default()
            fuente_negrita = ImageFont.load_default()

    y = 20
    draw.text((20, y), f"Territorio {territorio['numero']}", fill="black", font=fuente_titulo)
    y += 30
    draw.text((20, y), f"Estado: {territorio['estado']}", fill="black", font=fuente_normal)
    y += 40

    columnas = ["Dirección", "Teléfono", "No llamar", "Funciona"]
    x_pos = [20, 350, 550, 680]

    for texto, x in zip(columnas, x_pos):
        draw.text((x, y), texto, fill="black", font=fuente_negrita)
    y += fila_alto
    draw.line((20, y - 5, ancho_img - 20, y - 5), fill="gray")

    for reg in registros:
        if reg["funcionan"] is None:
            funciona_txt = "Sin verificar"
        elif reg["funcionan"]:
            funciona_txt = "Si"
        else:
            funciona_txt = "No"

        valores = [
            (reg["direccion"] or "—")[:35],
            reg["telefono"],
            "Si" if reg["no_llamar"] else "No",
            funciona_txt,
        ]
        for texto, x in zip(valores, x_pos):
            draw.text((x, y), str(texto), fill="black", font=fuente_normal)
        y += fila_alto

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Importar Excel de Historial de Asignaciones
# ---------------------------------------------------------------------------

# Columnas obligatorias del Excel de Historial (nombres exactos, sensibles a
# mayúsculas porque así vienen del Excel original).
COLUMNAS_HISTORIAL = ["Territorio", "Responsable", "Fecha asignado", "Fecha completado", "Detalles"]


def _celda_pandas_a_texto(valor):
    """
    Como _celda_a_texto, pero para valores que vienen de un DataFrame de pandas.
    Cubre dos casos típicos de leer Excel con pandas:
    - Celdas vacías, que pandas representa como NaN (float), no como None.
    - Columnas de números (como 'Territorio') que, si tienen alguna celda
      vacía en el medio, pandas infiere como float: 404 llega como 404.0.
    """
    if valor is None or pd.isna(valor):
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    return texto or None


def _parsear_fecha_excel(texto):
    """
    Convierte el texto de una celda de fecha a 'YYYY-MM-DD HH:MM:SS'
    (el mismo formato que usa el resto de la app), o None si la celda
    está vacía (incluyendo NaN de pandas) o no se pudo interpretar.

    Acepta 'YYYY-MM-DD' y 'DD/MM/YYYY', con o sin hora incluida (esto último
    pasa seguido porque Excel suele guardar las fechas como fecha+hora aunque
    la celda solo muestre el día).
    """
    if texto is None or pd.isna(texto):
        return None
    texto = str(texto).strip()
    if not texto:
        return None

    formatos = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y")
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    return None  # formato no reconocido; se trata como fecha vacía


def importar_historial(archivo, conn):
    """
    Lee el Excel de Historial de Asignaciones (columnas: Territorio,
    Responsable, Fecha asignado, Fecha completado, Detalles) y:

    1. Crea los responsables que no existan todavía (activo = 1).
    2. Crea los territorios que no existan todavía (por 'numero').
    3. Normaliza las fechas de cada fila.
    4. Hace upsert en 'asignaciones': la clave natural para no duplicar si el
       mismo Excel se sube dos veces es (territorio_id, responsable_id,
       fecha_asignado). Si ya existe esa asignación, se actualizan
       'fecha_finalizacion' y 'detalles' en vez de crear una fila nueva.
    5. Al final, sincroniza 'territorios.estado': queda 'En trabajo' si el
       territorio tiene alguna asignación sin fecha_finalizacion, o
       'Disponible' si no tiene ninguna abierta.

    Devuelve un diccionario con contadores para mostrarle un resumen al usuario.
    """
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        raise ValueError(f"No se pudo leer el Excel: {e}")

    faltantes = [c for c in COLUMNAS_HISTORIAL if c not in df.columns]
    if faltantes:
        raise ValueError(
            "El Excel debe tener las columnas: "
            f"{', '.join(COLUMNAS_HISTORIAL)}. Faltan: {', '.join(faltantes)}."
        )

    resumen = {
        "responsables_creados": 0,
        "territorios_creados": 0,
        "asignaciones_insertadas": 0,
        "asignaciones_actualizadas": 0,
        "errores": 0,
        "detalle_errores": [],
    }

    # Cachés en memoria para no repetir SELECTs por cada fila del Excel.
    cache_responsables = {}
    cache_territorios = {}
    territorios_afectados = set()

    for pos, fila in df.iterrows():
        num_fila_excel = pos + 2  # +2: índice 0-based de pandas + fila de encabezado

        numero = _celda_pandas_a_texto(fila["Territorio"])
        nombre_resp = _celda_pandas_a_texto(fila["Responsable"])
        fecha_asignado = _parsear_fecha_excel(fila["Fecha asignado"])
        fecha_completado = _parsear_fecha_excel(fila["Fecha completado"])
        detalles = _celda_pandas_a_texto(fila["Detalles"])

        if not numero or not nombre_resp or not fecha_asignado:
            resumen["errores"] += 1
            resumen["detalle_errores"].append(
                f"Fila {num_fila_excel}: faltan datos obligatorios "
                "(Territorio, Responsable o Fecha asignado válida), se saltea."
            )
            continue

        # --- Responsable: auto-crear si no existe ---
        responsable_id = cache_responsables.get(nombre_resp)
        if responsable_id is None:
            fila_resp = conn.execute(
                "SELECT id FROM responsables WHERE nombre = ?", (nombre_resp,)
            ).fetchone()
            if fila_resp is None:
                cur = conn.execute(
                    "INSERT INTO responsables (nombre, activo) VALUES (?, 1)",
                    (nombre_resp,),
                )
                responsable_id = cur.lastrowid
                resumen["responsables_creados"] += 1
            else:
                responsable_id = fila_resp["id"]
            cache_responsables[nombre_resp] = responsable_id

        # --- Territorio: auto-crear si no existe ---
        territorio_id = cache_territorios.get(numero)
        if territorio_id is None:
            fila_terr = conn.execute(
                "SELECT id FROM territorios WHERE numero = ?", (numero,)
            ).fetchone()
            if fila_terr is None:
                cur = conn.execute(
                    "INSERT INTO territorios (numero, estado) VALUES (?, 'Disponible')",
                    (numero,),
                )
                territorio_id = cur.lastrowid
                resumen["territorios_creados"] += 1
            else:
                territorio_id = fila_terr["id"]
            cache_territorios[numero] = territorio_id

        territorios_afectados.add(territorio_id)

        # --- Upsert de la asignación ---
        # Misma clave natural (territorio + responsable + fecha_asignado) =
        # se considera la misma asignación, aunque se reimporte el archivo.
        existente = conn.execute(
            """
            SELECT id FROM asignaciones
            WHERE territorio_id = ? AND responsable_id = ? AND fecha_asignado = ?
            """,
            (territorio_id, responsable_id, fecha_asignado),
        ).fetchone()

        if existente is None:
            conn.execute(
                """
                INSERT INTO asignaciones
                    (territorio_id, responsable_id, fecha_asignado, fecha_finalizacion, detalles)
                VALUES (?, ?, ?, ?, ?)
                """,
                (territorio_id, responsable_id, fecha_asignado, fecha_completado, detalles),
            )
            resumen["asignaciones_insertadas"] += 1
        else:
            conn.execute(
                "UPDATE asignaciones SET fecha_finalizacion = ?, detalles = ? WHERE id = ?",
                (fecha_completado, detalles, existente["id"]),
            )
            resumen["asignaciones_actualizadas"] += 1

    # --- Sincronizamos el estado de cada territorio que tocamos ---
    for territorio_id in territorios_afectados:
        abierta = conn.execute(
            "SELECT 1 FROM asignaciones WHERE territorio_id = ? AND fecha_finalizacion IS NULL",
            (territorio_id,),
        ).fetchone()
        nuevo_estado = "En trabajo" if abierta else "Disponible"
        conn.execute(
            "UPDATE territorios SET estado = ? WHERE id = ?",
            (nuevo_estado, territorio_id),
        )

    conn.commit()
    return resumen
