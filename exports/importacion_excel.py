"""
importacion_excel.py
---------------------
Importa (upsert) territorios y registros desde un Excel de números
telefónicos: columnas numero, direccion, telefono, observaciones,
no_llamar, funcionan, notas_internas.
"""

from openpyxl import load_workbook

from .celdas import _celda_a_texto, _celda_a_bool01, _celda_a_funcionan

# Columnas esperadas en la primera fila (encabezado) del Excel.
# El orden no importa, se buscan por nombre.
COLUMNAS_ESPERADAS = [
    "numero", "direccion", "telefono", "observaciones",
    "no_llamar", "funcionan", "notas_internas",
]


def importar_excel(archivo, conn):
    """
    Lee un archivo Excel (file-like) y hace upsert en 'territorios' y 'registros'.

    - 'numero' y 'telefono' son obligatorios en cada fila; si faltan, la fila
      se cuenta como error y se saltea (no frena la importación completa).
    - El territorio se crea si no existe (por 'numero', que es UNIQUE).
    - El registro se inserta o actualiza según la clave única
      (territorio_id, direccion, telefono) que ya define el esquema.

    Devuelve un diccionario con contadores para mostrarle un resumen al usuario.
    """
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active

    filas = list(hoja.iter_rows(values_only=False))
    if not filas:
        return {"territorios_creados": 0, "registros_insertados": 0,
                "registros_actualizados": 0, "errores": 0, "detalle_errores": []}

    # Mapeamos nombre de columna -> índice, leyendo la fila de encabezado
    encabezado = [str(c.value).strip().lower() if c.value else "" for c in filas[0]]
    indice = {}
    for nombre in COLUMNAS_ESPERADAS:
        if nombre in encabezado:
            indice[nombre] = encabezado.index(nombre)

    if "numero" not in indice or "telefono" not in indice:
        raise ValueError(
            "El Excel debe tener al menos las columnas 'numero' y 'telefono' "
            "en la primera fila."
        )

    resumen = {
        "territorios_creados": 0,
        "registros_insertados": 0,
        "registros_actualizados": 0,
        "errores": 0,
        "detalle_errores": [],
    }

    def valor(fila, nombre):
        i = indice.get(nombre)
        return fila[i].value if i is not None and i < len(fila) else None

    for num_fila, fila in enumerate(filas[1:], start=2):  # fila 1 es encabezado
        numero = _celda_a_texto(valor(fila, "numero"))
        telefono = _celda_a_texto(valor(fila, "telefono"))

        if not numero or not telefono:
            resumen["errores"] += 1
            resumen["detalle_errores"].append(
                f"Fila {num_fila}: falta 'numero' o 'telefono', se saltea."
            )
            continue

        direccion = _celda_a_texto(valor(fila, "direccion"))
        observaciones = _celda_a_texto(valor(fila, "observaciones"))
        notas_internas = _celda_a_texto(valor(fila, "notas_internas"))
        no_llamar = _celda_a_bool01(valor(fila, "no_llamar"))
        funcionan = _celda_a_funcionan(valor(fila, "funcionan"))

        # --- Upsert de territorio por 'numero' ---
        territorio = conn.execute(
            "SELECT id FROM territorios WHERE numero = ?", (numero,)
        ).fetchone()
        if territorio is None:
            cur = conn.execute(
                "INSERT INTO territorios (numero, estado) VALUES (?, 'Disponible')",
                (numero,),
            )
            territorio_id = cur.lastrowid
            resumen["territorios_creados"] += 1
        else:
            territorio_id = territorio["id"]

        # --- Upsert de registro por (territorio_id, direccion, telefono) ---
        existente = conn.execute(
            """
            SELECT id FROM registros
            WHERE territorio_id = ?
              AND direccion IS ?
              AND telefono = ?
            """,
            (territorio_id, direccion, telefono),
        ).fetchone()

        if existente is None:
            conn.execute(
                """
                INSERT INTO registros
                    (territorio_id, direccion, telefono, observaciones,
                     no_llamar, funcionan, notas_internas)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (territorio_id, direccion, telefono, observaciones,
                 no_llamar, funcionan, notas_internas),
            )
            resumen["registros_insertados"] += 1
        else:
            conn.execute(
                """
                UPDATE registros
                SET observaciones = ?, no_llamar = ?, funcionan = ?, notas_internas = ?
                WHERE id = ?
                """,
                (observaciones, no_llamar, funcionan, notas_internas, existente["id"]),
            )
            resumen["registros_actualizados"] += 1

    conn.commit()
    return resumen
