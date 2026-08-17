"""
exportacion_excel.py
---------------------
Exporta una lista de registros (con su numero_territorio) a un .xlsx con
encabezado en color, bordes prolijos y filas alternadas.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .paleta import HEADER_HEX, FILA_PAR_HEX, BORDE_HEX, TEXTO_OSCURO_HEX


def generar_excel(registros):
    """
    Recibe una lista de filas (sqlite3.Row) con columnas de 'registros' +
    'numero_territorio', y devuelve un BytesIO con un .xlsx listo para descargar.

    El resultado sigue la misma estética que las planillas en papel: encabezado
    en color con texto en negrita, bordes prolijos en cada celda y filas
    alternadas para que sea fácil de leer con muchos registros.
    """
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Registros"

    # Nombres de columna + etiquetas legibles para mostrar (la clave de la
    # izquierda es la que se usa para leer cada fila de 'registros').
    columnas = [
        ("numero_territorio", "N° Territorio"),
        ("direccion", "Dirección"),
        ("telefono", "Teléfono"),
        ("observaciones", "Observaciones"),
        ("no_llamar", "No llamar"),
        ("funcionan", "Funciona"),
        ("notas_internas", "Notas internas"),
    ]
    encabezados = [etiqueta for _, etiqueta in columnas]
    hoja.append(encabezados)

    # --- Paleta y estilos reutilizables ---
    relleno_header = PatternFill(start_color=HEADER_HEX, end_color=HEADER_HEX, fill_type="solid")
    relleno_par = PatternFill(start_color=FILA_PAR_HEX, end_color=FILA_PAR_HEX, fill_type="solid")
    fuente_header = Font(bold=True, size=11, color=TEXTO_OSCURO_HEX)
    alineacion_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_texto = Alignment(horizontal="left", vertical="center", wrap_text=True)

    borde_fino = Side(style="thin", color=BORDE_HEX)
    borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    # Columnas que van centradas por ser valores cortos (territorio, sí/no)
    columnas_centradas = {"numero_territorio", "no_llamar", "funcionan"}

    # --- Encabezado ---
    for celda in hoja[1]:
        celda.fill = relleno_header
        celda.font = fuente_header
        celda.alignment = alineacion_header
        celda.border = borde_celda
    hoja.row_dimensions[1].height = 22

    # --- Filas de datos ---
    fila_actual = 2
    for reg in registros:
        funcionan = reg["funcionan"]
        funcionan_texto = "" if funcionan is None else ("Sí" if funcionan else "No")
        hoja.append([
            reg["numero_territorio"],
            reg["direccion"] or "",
            reg["telefono"],
            reg["observaciones"] or "",
            "Sí" if reg["no_llamar"] else "No",
            funcionan_texto,
            reg["notas_internas"] or "",
        ])

        es_par = (fila_actual % 2 == 0)
        for col_idx, (clave, _) in enumerate(columnas, start=1):
            celda = hoja.cell(row=fila_actual, column=col_idx)
            celda.border = borde_celda
            celda.alignment = alineacion_centro if clave in columnas_centradas else alineacion_texto
            if es_par:
                celda.fill = relleno_par

        fila_actual += 1

    # Ancho de columnas prolijo, para que no quede todo apretado
    anchos = [14, 30, 16, 32, 12, 12, 30]
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    # Encabezado siempre visible al scrollear + filtro rápido por columna
    hoja.freeze_panes = "A2"
    if fila_actual > 2:
        hoja.auto_filter.ref = hoja.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
