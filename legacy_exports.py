"""
exports.py
----------
Funciones auxiliares para:
- Importar territorios/registros desde un archivo Excel (upsert)
- Exportar registros a Excel
- Exportar un territorio a PDF
- Exportar un territorio a PNG

Cada función devuelve un objeto BytesIO listo para mandar con send_file,
excepto importar_excel que devuelve un resumen (dict) de lo que hizo.
"""

import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from PIL import Image, ImageDraw, ImageFont


# ---------------------------------------------------------------------------
# Helpers para interpretar celdas de Excel de forma flexible
# ---------------------------------------------------------------------------

def _celda_a_texto(valor):
    """Convierte una celda de Excel a texto limpio, o None si está vacía."""
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _celda_a_bool01(valor):
    """
    Convierte una celda a 0 o 1, aceptando varias formas de escribir "sí/no":
    1, 0, "si", "sí", "no", "true", "false", vacío (-> 0).
    """
    if valor is None:
        return 0
    texto = str(valor).strip().lower()
    if texto in ("1", "si", "sí", "true", "x"):
        return 1
    return 0


def _celda_a_funcionan(valor):
    """
    Convierte una celda al valor de 'funcionan': None (sin verificar), 1 o 0.
    Deja en None cualquier celda vacía o que no se entienda con claridad.
    """
    if valor is None:
        return None
    texto = str(valor).strip().lower()
    if texto in ("1", "si", "sí", "true"):
        return 1
    if texto in ("0", "no", "false"):
        return 0
    return None


# ---------------------------------------------------------------------------
# Importar Excel (upsert de territorios + registros)
# ---------------------------------------------------------------------------

# Columnas esperadas en la primera fila (encabezado) del Excel.
# El orden no importa, se buscan por nombre.
COLUMNAS_ESPERADAS = [
    "numero", "direccion", "telefono", "observaciones",
    "no_llamar", "funcionan", "notas_internas",
]


def importar_excel(archivo, conn):
    """
    Lee un archivo Excel (file-like) y hace upsert en 'territorios' y 'registros'.

    - 'numero' y 'telefono' son obligatorios en cada fila; si faltan, la fila
      se cuenta como error y se saltea (no frena la importación completa).
    - El territorio se crea si no existe (por 'numero', que es UNIQUE).
    - El registro se inserta o actualiza según la clave única
      (territorio_id, direccion, telefono) que ya define el esquema.

    Devuelve un diccionario con contadores para mostrarle un resumen al usuario.
    """
    wb = load_workbook(archivo, data_only=True)
    hoja = wb.active

    filas = list(hoja.iter_rows(values_only=False))
    if not filas:
        return {"territorios_creados": 0, "registros_insertados": 0,
                "registros_actualizados": 0, "errores": 0, "detalle_errores": []}

    # Mapeamos nombre de columna -> índice, leyendo la fila de encabezado
    encabezado = [str(c.value).strip().lower() if c.value else "" for c in filas[0]]
    indice = {}
    for nombre in COLUMNAS_ESPERADAS:
        if nombre in encabezado:
            indice[nombre] = encabezado.index(nombre)

    if "numero" not in indice or "telefono" not in indice:
        raise ValueError(
            "El Excel debe tener al menos las columnas 'numero' y 'telefono' "
            "en la primera fila."
        )

    resumen = {
        "territorios_creados": 0,
        "registros_insertados": 0,
        "registros_actualizados": 0,
        "errores": 0,
        "detalle_errores": [],
    }

    def valor(fila, nombre):
        i = indice.get(nombre)
        return fila[i].value if i is not None and i < len(fila) else None

    for num_fila, fila in enumerate(filas[1:], start=2):  # fila 1 es encabezado
        numero = _celda_a_texto(valor(fila, "numero"))
        telefono = _celda_a_texto(valor(fila, "telefono"))

        if not numero or not telefono:
            resumen["errores"] += 1
            resumen["detalle_errores"].append(
                f"Fila {num_fila}: falta 'numero' o 'telefono', se saltea."
            )
            continue

        direccion = _celda_a_texto(valor(fila, "direccion"))
        observaciones = _celda_a_texto(valor(fila, "observaciones"))
        notas_internas = _celda_a_texto(valor(fila, "notas_internas"))
        no_llamar = _celda_a_bool01(valor(fila, "no_llamar"))
        funcionan = _celda_a_funcionan(valor(fila, "funcionan"))

        # --- Upsert de territorio por 'numero' ---
        territorio = conn.execute(
            "SELECT id FROM territorios WHERE numero = ?", (numero,)
        ).fetchone()
        if territorio is None:
            cur = conn.execute(
                "INSERT INTO territorios (numero, estado) VALUES (?, 'Disponible')",
                (numero,),
            )
            territorio_id = cur.lastrowid
            resumen["territorios_creados"] += 1
        else:
            territorio_id = territorio["id"]

        # --- Upsert de registro por (territorio_id, direccion, telefono) ---
        existente = conn.execute(
            """
            SELECT id FROM registros
            WHERE territorio_id = ?
              AND direccion IS ?
              AND telefono = ?
            """,
            (territorio_id, direccion, telefono),
        ).fetchone()

        if existente is None:
            conn.execute(
                """
                INSERT INTO registros
                    (territorio_id, direccion, telefono, observaciones,
                     no_llamar, funcionan, notas_internas)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (territorio_id, direccion, telefono, observaciones,
                 no_llamar, funcionan, notas_internas),
            )
            resumen["registros_insertados"] += 1
        else:
            conn.execute(
                """
                UPDATE registros
                SET observaciones = ?, no_llamar = ?, funcionan = ?, notas_internas = ?
                WHERE id = ?
                """,
                (observaciones, no_llamar, funcionan, notas_internas, existente["id"]),
            )
            resumen["registros_actualizados"] += 1

    conn.commit()
    return resumen


# ---------------------------------------------------------------------------
# Exportar a Excel
# ---------------------------------------------------------------------------

def generar_excel(registros):
    """
    Recibe una lista de filas (sqlite3.Row) con columnas de 'registros' +
    'numero_territorio', y devuelve un BytesIO con un .xlsx listo para descargar.

    El resultado sigue la misma estética que las planillas en papel: encabezado
    en color con texto en negrita, bordes prolijos en cada celda y filas
    alternadas para que sea fácil de leer con muchos registros.
    """
    wb = Workbook()
    hoja = wb.active
    hoja.title = "Registros"

    # Nombres de columna + etiquetas legibles para mostrar (la clave de la
    # izquierda es la que se usa para leer cada fila de 'registros').
    columnas = [
        ("numero_territorio", "N° Territorio"),
        ("direccion", "Dirección"),
        ("telefono", "Teléfono"),
        ("observaciones", "Observaciones"),
        ("no_llamar", "No llamar"),
        ("funcionan", "Funciona"),
        ("notas_internas", "Notas internas"),
    ]
    encabezados = [etiqueta for _, etiqueta in columnas]
    hoja.append(encabezados)

    # --- Paleta y estilos reutilizables ---
    COLOR_HEADER = "F4B183"   # naranja suave, mismo espíritu que las planillas impresas
    COLOR_FILA_PAR = "FBEEE6"  # banda muy tenue para alternar filas

    relleno_header = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    relleno_par = PatternFill(start_color=COLOR_FILA_PAR, end_color=COLOR_FILA_PAR, fill_type="solid")
    fuente_header = Font(bold=True, size=11, color="3B2A1A")
    alineacion_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alineacion_centro = Alignment(horizontal="center", vertical="center")
    alineacion_texto = Alignment(horizontal="left", vertical="center", wrap_text=True)

    borde_fino = Side(style="thin", color="BFBFBF")
    borde_celda = Border(left=borde_fino, right=borde_fino, top=borde_fino, bottom=borde_fino)

    # Columnas que van centradas por ser valores cortos (territorio, sí/no)
    columnas_centradas = {"numero_territorio", "no_llamar", "funcionan"}

    # --- Encabezado ---
    for celda in hoja[1]:
        celda.fill = relleno_header
        celda.font = fuente_header
        celda.alignment = alineacion_header
        celda.border = borde_celda
    hoja.row_dimensions[1].height = 22

    # --- Filas de datos ---
    fila_actual = 2
    for reg in registros:
        funcionan = reg["funcionan"]
        funcionan_texto = "" if funcionan is None else ("Sí" if funcionan else "No")
        hoja.append([
            reg["numero_territorio"],
            reg["direccion"] or "",
            reg["telefono"],
            reg["observaciones"] or "",
            "Sí" if reg["no_llamar"] else "No",
            funcionan_texto,
            reg["notas_internas"] or "",
        ])

        es_par = (fila_actual % 2 == 0)
        for col_idx, (clave, _) in enumerate(columnas, start=1):
            celda = hoja.cell(row=fila_actual, column=col_idx)
            celda.border = borde_celda
            celda.alignment = alineacion_centro if clave in columnas_centradas else alineacion_texto
            if es_par:
                celda.fill = relleno_par

        fila_actual += 1

    # Ancho de columnas prolijo, para que no quede todo apretado
    anchos = [14, 30, 16, 32, 12, 12, 30]
    for i, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    # Encabezado siempre visible al scrollear + filtro rápido por columna
    hoja.freeze_panes = "A2"
    if fila_actual > 2:
        hoja.auto_filter.ref = hoja.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Exportar territorio a PDF
# ---------------------------------------------------------------------------

def generar_pdf(territorio, registros):
    """
    Genera un PDF prolijo e imprimible con los datos del territorio y su
    tabla de registros telefónicos: encabezado en color, bordes en toda la
    tabla y filas alternadas (misma estética que el Excel exportado).

    'No llamar' y 'Funciona' quedan en blanco cuando no hay un dato cargado
    (en vez de escribir "No" / "Sin verificar"), para no confundir y para
    que sea cómodo completarlas a mano si se imprime la planilla.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    COLOR_HEADER = colors.HexColor("#F4B183")
    COLOR_FILA_PAR = colors.HexColor("#FBEEE6")
    COLOR_BORDE = colors.HexColor("#BFBFBF")
    COLOR_TEXTO_OSCURO = colors.HexColor("#3B2A1A")

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloTerritorio", parent=estilos["Heading1"],
        fontSize=18, spaceAfter=2, textColor=COLOR_TEXTO_OSCURO,
    )
    estilo_subtitulo = ParagraphStyle(
        "Subtitulo", parent=estilos["Normal"],
        fontSize=11, textColor=colors.HexColor("#555555"), spaceAfter=14,
    )
    estilo_celda = ParagraphStyle("Celda", parent=estilos["Normal"], fontSize=9, leading=11)
    estilo_celda_centro = ParagraphStyle("CeldaCentro", parent=estilo_celda, alignment=TA_CENTER)
    estilo_encabezado_celda = ParagraphStyle(
        "EncabezadoCelda", parent=estilos["Normal"], fontSize=9.5, leading=12,
        textColor=COLOR_TEXTO_OSCURO, alignment=TA_CENTER, fontName="Helvetica-Bold",
    )

    elementos = [
        Paragraph(f"Territorio N.° {territorio['numero']}", estilo_titulo),
        Paragraph(f"Estado: {territorio['estado']}", estilo_subtitulo),
    ]

    encabezados = ["Dirección", "Teléfono", "Observaciones", "No llamar", "Funciona"]
    filas = [[Paragraph(h, estilo_encabezado_celda) for h in encabezados]]

    for reg in registros:
        # En blanco si no hay dato cargado, para no escribir "No" o
        # "Sin verificar" en cada fila y que sea fácil completar a mano.
        no_llamar_txt = "Sí" if reg["no_llamar"] else ""
        if reg["funcionan"] is None:
            funciona_txt = ""
        else:
            funciona_txt = "Sí" if reg["funcionan"] else "No"

        filas.append([
            Paragraph(reg["direccion"] or "—", estilo_celda),
            Paragraph(reg["telefono"] or "", estilo_celda),
            Paragraph(reg["observaciones"] or "", estilo_celda),
            Paragraph(no_llamar_txt, estilo_celda_centro),
            Paragraph(funciona_txt, estilo_celda_centro),
        ])

    anchos_col = [5.3 * cm, 2.8 * cm, 5.4 * cm, 2.3 * cm, 2.3 * cm]
    tabla = Table(filas, colWidths=anchos_col, repeatRows=1)

    estilo_tabla = [
        ("BACKGROUND", (0, 0), (-1, 0), COLOR_HEADER),
        ("GRID", (0, 0), (-1, -1), 0.6, COLOR_BORDE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    for i in range(1, len(filas)):
        if i % 2 == 0:
            estilo_tabla.append(("BACKGROUND", (0, i), (-1, i), COLOR_FILA_PAR))
    tabla.setStyle(TableStyle(estilo_tabla))

    elementos.append(tabla)
    doc.build(elementos)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Exportar territorio a PNG
# ---------------------------------------------------------------------------

def generar_png(territorio, registros):
    """
    Genera una imagen PNG con una tabla del territorio, pensada para
    compartir rápido (ej. por WhatsApp) sin necesidad de abrir un PDF.

    Misma estética que el Excel/PDF exportados: encabezado en color y grilla
    de bordes prolija. 'No llamar' y 'Funciona' quedan en blanco cuando no
    hay un dato cargado, en vez de escribir "No" / "Sin verificar".
    """
    ancho_img = 900
    margen_izq = 20
    margen_der = ancho_img - 20
    fila_alto = 32
    alto_titulo = 78
    alto_encabezado_tabla = 34
    alto_img = alto_titulo + alto_encabezado_tabla + fila_alto * max(len(registros), 1) + 20

    img = Image.new("RGB", (ancho_img, alto_img), color="white")
    draw = ImageDraw.Draw(img)

    try:
        fuente_titulo = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20
        )
        fuente_normal = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14
        )
        fuente_negrita = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14
        )
    except OSError:
        # Intentar con fuentes de Windows
        try:
            fuente_titulo = ImageFont.truetype("arial.ttf", 20)
            fuente_normal = ImageFont.truetype("arial.ttf", 14)
            fuente_negrita = ImageFont.truetype("arialbd.ttf", 14)
        except OSError:
            # Último recurso: fuente por defecto de Pillow
            fuente_titulo = ImageFont.load_default()
            fuente_normal = ImageFont.load_default()
            fuente_negrita = ImageFont.load_default()

    COLOR_HEADER = (244, 177, 131)     # F4B183
    COLOR_FILA_PAR = (251, 238, 230)   # FBEEE6
    COLOR_BORDE = (191, 191, 191)      # BFBFBF
    COLOR_TEXTO_HEADER = (59, 42, 26)  # 3B2A1A

    y = 20
    draw.text((margen_izq, y), f"Territorio {territorio['numero']}", fill="black", font=fuente_titulo)
    y += 30
    draw.text((margen_izq, y), f"Estado: {territorio['estado']}", fill="black", font=fuente_normal)
    y += 34

    columnas = ["Dirección", "Teléfono", "No llamar", "Funciona"]
    anchos_col = [340, 200, 150, 170]
    x_bordes = [margen_izq]
    for ancho_col in anchos_col:
        x_bordes.append(x_bordes[-1] + ancho_col)

    y_tabla_inicio = y

    # Encabezado con fondo de color
    draw.rectangle([margen_izq, y, x_bordes[-1], y + alto_encabezado_tabla], fill=COLOR_HEADER)
    for texto, x_izq in zip(columnas, x_bordes[:-1]):
        draw.text((x_izq + 8, y + 9), texto, fill=COLOR_TEXTO_HEADER, font=fuente_negrita)
    y += alto_encabezado_tabla

    for idx, reg in enumerate(registros):
        if idx % 2 == 1:
            draw.rectangle([margen_izq, y, x_bordes[-1], y + fila_alto], fill=COLOR_FILA_PAR)

        # En blanco si no hay dato cargado, para no escribir "No" o
        # "Sin verificar" en cada fila y que sea fácil completar a mano.
        no_llamar_txt = "Sí" if reg["no_llamar"] else ""
        if reg["funcionan"] is None:
            funciona_txt = ""
        else:
            funciona_txt = "Sí" if reg["funcionan"] else "No"

        valores = [
            (reg["direccion"] or "—")[:45],
            reg["telefono"] or "",
            no_llamar_txt,
            funciona_txt,
        ]
        for texto, x_izq in zip(valores, x_bordes[:-1]):
            draw.text((x_izq + 8, y + 9), str(texto), fill="black", font=fuente_normal)
        y += fila_alto

    y_tabla_fin = y

    # Grilla: líneas verticales entre columnas + horizontales entre filas
    for x in x_bordes:
        draw.line([(x, y_tabla_inicio), (x, y_tabla_fin)], fill=COLOR_BORDE, width=1)
    y_linea = y_tabla_inicio
    draw.line([(margen_izq, y_linea), (x_bordes[-1], y_linea)], fill=COLOR_BORDE, width=1)
    y_linea += alto_encabezado_tabla
    draw.line([(margen_izq, y_linea), (x_bordes[-1], y_linea)], fill=COLOR_BORDE, width=1)
    for _ in registros:
        y_linea += fila_alto
        draw.line([(margen_izq, y_linea), (x_bordes[-1], y_linea)], fill=COLOR_BORDE, width=1)

    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Importar Excel de Historial de Asignaciones
# ---------------------------------------------------------------------------

# Columnas obligatorias del Excel de Historial (nombres exactos, sensibles a
# mayúsculas porque así vienen del Excel original).
COLUMNAS_HISTORIAL = ["Territorio", "Responsable", "Fecha asignado", "Fecha completado", "Detalles"]


def _celda_pandas_a_texto(valor):
    """
    Como _celda_a_texto, pero para valores que vienen de un DataFrame de pandas.
    Cubre dos casos típicos de leer Excel con pandas:
    - Celdas vacías, que pandas representa como NaN (float), no como None.
    - Columnas de números (como 'Territorio') que, si tienen alguna celda
      vacía en el medio, pandas infiere como float: 404 llega como 404.0.
    """
    if valor is None or pd.isna(valor):
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip()
    return texto or None


def _parsear_fecha_excel(texto):
    """
    Convierte el texto de una celda de fecha a 'YYYY-MM-DD HH:MM:SS'
    (el mismo formato que usa el resto de la app), o None si la celda
    está vacía (incluyendo NaN de pandas) o no se pudo interpretar.

    Acepta 'YYYY-MM-DD' y 'DD/MM/YYYY', con o sin hora incluida (esto último
    pasa seguido porque Excel suele guardar las fechas como fecha+hora aunque
    la celda solo muestre el día).
    """
    if texto is None or pd.isna(texto):
        return None
    texto = str(texto).strip()
    if not texto:
        return None

    formatos = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y")
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue

    return None  # formato no reconocido; se trata como fecha vacía


def importar_historial(archivo, conn):
    """
    Lee el Excel de Historial de Asignaciones (columnas: Territorio,
    Responsable, Fecha asignado, Fecha completado, Detalles) y:

    1. Crea los responsables que no existan todavía (activo = 1).
    2. Crea los territorios que no existan todavía (por 'numero').
    3. Normaliza las fechas de cada fila.
    4. Hace upsert en 'asignaciones': la clave natural para no duplicar si el
       mismo Excel se sube dos veces es (territorio_id, responsable_id,
       fecha_asignado). Si ya existe esa asignación, se actualizan
       'fecha_finalizacion' y 'detalles' en vez de crear una fila nueva.
    5. Al final, sincroniza 'territorios.estado': queda 'En trabajo' si el
       territorio tiene alguna asignación sin fecha_finalizacion, o
       'Disponible' si no tiene ninguna abierta.

    Devuelve un diccionario con contadores para mostrarle un resumen al usuario.
    """
    try:
        df = pd.read_excel(archivo)
    except Exception as e:
        raise ValueError(f"No se pudo leer el Excel: {e}")

    faltantes = [c for c in COLUMNAS_HISTORIAL if c not in df.columns]
    if faltantes:
        raise ValueError(
            "El Excel debe tener las columnas: "
            f"{', '.join(COLUMNAS_HISTORIAL)}. Faltan: {', '.join(faltantes)}."
        )

    resumen = {
        "responsables_creados": 0,
        "territorios_creados": 0,
        "asignaciones_insertadas": 0,
        "asignaciones_actualizadas": 0,
        "errores": 0,
        "detalle_errores": [],
    }

    # Cachés en memoria para no repetir SELECTs por cada fila del Excel.
    cache_responsables = {}
    cache_territorios = {}
    territorios_afectados = set()

    for pos, fila in df.iterrows():
        num_fila_excel = pos + 2  # +2: índice 0-based de pandas + fila de encabezado

        numero = _celda_pandas_a_texto(fila["Territorio"])
        nombre_resp = _celda_pandas_a_texto(fila["Responsable"])
        fecha_asignado = _parsear_fecha_excel(fila["Fecha asignado"])
        fecha_completado = _parsear_fecha_excel(fila["Fecha completado"])
        detalles = _celda_pandas_a_texto(fila["Detalles"])

        if not numero or not nombre_resp or not fecha_asignado:
            resumen["errores"] += 1
            resumen["detalle_errores"].append(
                f"Fila {num_fila_excel}: faltan datos obligatorios "
                "(Territorio, Responsable o Fecha asignado válida), se saltea."
            )
            continue

        # --- Responsable: auto-crear si no existe ---
        responsable_id = cache_responsables.get(nombre_resp)
        if responsable_id is None:
            fila_resp = conn.execute(
                "SELECT id FROM responsables WHERE nombre = ?", (nombre_resp,)
            ).fetchone()
            if fila_resp is None:
                cur = conn.execute(
                    "INSERT INTO responsables (nombre, activo) VALUES (?, 1)",
                    (nombre_resp,),
                )
                responsable_id = cur.lastrowid
                resumen["responsables_creados"] += 1
            else:
                responsable_id = fila_resp["id"]
            cache_responsables[nombre_resp] = responsable_id

        # --- Territorio: auto-crear si no existe ---
        territorio_id = cache_territorios.get(numero)
        if territorio_id is None:
            fila_terr = conn.execute(
                "SELECT id FROM territorios WHERE numero = ?", (numero,)
            ).fetchone()
            if fila_terr is None:
                cur = conn.execute(
                    "INSERT INTO territorios (numero, estado) VALUES (?, 'Disponible')",
                    (numero,),
                )
                territorio_id = cur.lastrowid
                resumen["territorios_creados"] += 1
            else:
                territorio_id = fila_terr["id"]
            cache_territorios[numero] = territorio_id

        territorios_afectados.add(territorio_id)

        # --- Upsert de la asignación ---
        # Misma clave natural (territorio + responsable + fecha_asignado) =
        # se considera la misma asignación, aunque se reimporte el archivo.
        existente = conn.execute(
            """
            SELECT id FROM asignaciones
            WHERE territorio_id = ? AND responsable_id = ? AND fecha_asignado = ?
            """,
            (territorio_id, responsable_id, fecha_asignado),
        ).fetchone()

        if existente is None:
            conn.execute(
                """
                INSERT INTO asignaciones
                    (territorio_id, responsable_id, fecha_asignado, fecha_finalizacion, detalles)
                VALUES (?, ?, ?, ?, ?)
                """,
                (territorio_id, responsable_id, fecha_asignado, fecha_completado, detalles),
            )
            resumen["asignaciones_insertadas"] += 1
        else:
            conn.execute(
                "UPDATE asignaciones SET fecha_finalizacion = ?, detalles = ? WHERE id = ?",
                (fecha_completado, detalles, existente["id"]),
            )
            resumen["asignaciones_actualizadas"] += 1

    # --- Sincronizamos el estado de cada territorio que tocamos ---
    for territorio_id in territorios_afectados:
        abierta = conn.execute(
            "SELECT 1 FROM asignaciones WHERE territorio_id = ? AND fecha_finalizacion IS NULL",
            (territorio_id,),
        ).fetchone()
        nuevo_estado = "En trabajo" if abierta else "Disponible"
        conn.execute(
            "UPDATE territorios SET estado = ? WHERE id = ?",
            (nuevo_estado, territorio_id),
        )

    conn.commit()
    return resumen